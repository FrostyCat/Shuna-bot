import asyncio
import io
import os
from datetime import datetime, UTC

import discord
from discord.ext import commands
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from sqlalchemy import or_

from coc_api import get_player, verify_player_token
from db import Session
from models import DiscordUser, Player, GuildConfig
from helpers import add_player_to_db

_NOTIFY_GUILD_ID = os.getenv("NOTIFY_GUILD_ID", "")


def _player_tag_choices(current: str, linked_only: bool):
    session = Session()
    try:
        query = session.query(Player)
        if linked_only:
            query = query.filter(Player.discord_user_id.isnot(None))
        if current:
            like = f"%{current}%"
            query = query.filter(or_(Player.tag.ilike(like), Player.name.ilike(like)))
        players = query.order_by(Player.name).limit(25).all()
        return [
            discord.OptionChoice(name=f"{p.name} ({p.tag})", value=p.tag)
            for p in players
        ]
    finally:
        session.close()


def _players_by_discord_id(session, discord_ids: list[str]) -> dict[str, list[Player]]:
    if not discord_ids:
        return {}
    rows = (
        session.query(DiscordUser.discord_id, Player)
        .join(Player, Player.discord_user_id == DiscordUser.id)
        .filter(DiscordUser.discord_id.in_(discord_ids))
        .all()
    )
    result: dict[str, list[Player]] = {}
    for discord_id, player in rows:
        result.setdefault(discord_id, []).append(player)
    return result


async def player_tag_autocomplete(ctx: discord.AutocompleteContext):
    return await asyncio.to_thread(_player_tag_choices, ctx.value.lower(), False)


async def linked_tag_autocomplete(ctx: discord.AutocompleteContext):
    return await asyncio.to_thread(_player_tag_choices, ctx.value.lower(), True)


async def _notify_new_player_link(bot, guild_id: int, name: str, tag: str):
    session = Session()
    try:
        config = session.query(GuildConfig).filter_by(guild_id=str(guild_id)).first()
        if not config or not config.log_channel_id:
            return
        if _NOTIFY_GUILD_ID and str(guild_id) != _NOTIFY_GUILD_ID:
            return
        ch = bot.get_channel(int(config.log_channel_id))
        if not ch:
            return
        embed = discord.Embed(
            title="New Player Tracking Started",
            description=(
                f"**{name}** (`{tag}`) has been added to the tracking system.\n"
                f"Stats collection starts now — first-day data will be skipped to ensure accuracy."
            ),
            color=0xf472b6,
        )
        await ch.send(embed=embed)
    except Exception as e:
        print(f"Notify error for {tag}: {e}")
    finally:
        session.close()


class LinkCog(discord.Cog):
    def __init__(self, bot: discord.Bot):
        self.bot = bot

    @discord.slash_command(name="link", description="Link a Clash of Clans account to Discord")
    async def link(
        self,
        ctx: discord.ApplicationContext,
        tag: discord.Option(str, "Player tag, e.g. #ABC123", autocomplete=player_tag_autocomplete),
        user: discord.Option(discord.Member, "Discord user"),
        api_token: discord.Option(str, "API token from in-game settings"),
    ):
        await ctx.defer(ephemeral=True)

        tag = tag.upper().replace("O", "0")
        if not tag.startswith("#"):
            tag = "#" + tag

        valid = await verify_player_token(tag, api_token)
        if not valid:
            await ctx.followup.send("❌ Invalid API token. Check the token in your in-game settings.", ephemeral=True)
            return

        session = Session()
        player = await asyncio.to_thread(session.query(Player).filter_by(tag=tag).first)
        is_new = False
        if not player:
            result = await add_player_to_db(tag, session)
            if not result["success"]:
                await ctx.followup.send("❌ " + result["error"], ephemeral=True)
                session.close()
                return
            is_new = result.get("is_new", False)
            player = await asyncio.to_thread(session.query(Player).filter_by(tag=result["tag"]).first)

        discord_user = await asyncio.to_thread(session.query(DiscordUser).filter_by(discord_id=str(user.id)).first)
        if not discord_user:
            discord_user = DiscordUser(discord_id=str(user.id))
            session.add(discord_user)
            await asyncio.to_thread(session.flush)

        if player.discord_user_id == discord_user.id:
            session.close()
            await ctx.followup.send(f"ℹ️ **{player.name}** is already linked to {user.mention}.", ephemeral=True)
            return

        player.discord_user_id = discord_user.id
        player.is_verified = True
        player.verified_at = datetime.now(UTC)
        player_name = player.name
        await asyncio.to_thread(session.commit)
        session.close()

        if is_new and ctx.guild:
            await _notify_new_player_link(self.bot, ctx.guild.id, player_name, tag)

        await ctx.followup.send(f"✅ Linked **{player_name}** ({tag}) to {user.mention}.", ephemeral=True)

    @discord.slash_command(
        name="force_link",
        description="Link a CoC account to Discord without token verification (admin only)",
        default_member_permissions=discord.Permissions(administrator=True),
    )
    async def force_link(
        self,
        ctx: discord.ApplicationContext,
        tag: discord.Option(str, "Player tag, e.g. #ABC123", autocomplete=player_tag_autocomplete),
        user: discord.Option(discord.Member, "Discord user"),
    ):
        await ctx.defer(ephemeral=True)

        tag = tag.upper().replace("O", "0")
        if not tag.startswith("#"):
            tag = "#" + tag

        session = Session()
        player = await asyncio.to_thread(session.query(Player).filter_by(tag=tag).first)
        is_new = False
        if not player:
            result = await add_player_to_db(tag, session)
            if not result["success"]:
                await ctx.followup.send("❌ " + result["error"], ephemeral=True)
                session.close()
                return
            is_new = result.get("is_new", False)
            player = await asyncio.to_thread(session.query(Player).filter_by(tag=result["tag"]).first)

        discord_user = await asyncio.to_thread(session.query(DiscordUser).filter_by(discord_id=str(user.id)).first)
        if not discord_user:
            discord_user = DiscordUser(discord_id=str(user.id))
            session.add(discord_user)
            await asyncio.to_thread(session.flush)

        if player.discord_user_id == discord_user.id:
            session.close()
            await ctx.followup.send(f"ℹ️ **{player.name}** is already linked to {user.mention}.", ephemeral=True)
            return

        player.discord_user_id = discord_user.id
        player_name = player.name
        await asyncio.to_thread(session.commit)
        session.close()

        if is_new and ctx.guild:
            await _notify_new_player_link(self.bot, ctx.guild.id, player_name, tag)

        await ctx.followup.send(f"✅ Linked **{player_name}** ({tag}) to {user.mention}.", ephemeral=True)

    @discord.slash_command(name="unlink", description="Unlink a Clash of Clans account from Discord")
    async def unlink(
        self,
        ctx: discord.ApplicationContext,
        tag: discord.Option(str, "Player tag, e.g. #ABC123", autocomplete=linked_tag_autocomplete),
        user: discord.Option(discord.Member, "Discord user"),
        api_token: discord.Option(str, "API token from in-game settings"),
    ):
        await ctx.defer(ephemeral=True)

        tag = tag.upper().replace("O", "0")
        if not tag.startswith("#"):
            tag = "#" + tag

        valid = await verify_player_token(tag, api_token)
        if not valid:
            await ctx.followup.send("❌ Invalid API token.", ephemeral=True)
            return

        session = Session()
        player = await asyncio.to_thread(session.query(Player).filter_by(tag=tag).first)

        if not player or not player.discord_user or player.discord_user.discord_id != str(user.id):
            session.close()
            await ctx.followup.send("❌ This account is not linked to the given user.", ephemeral=True)
            return

        player_name = player.name
        player.discord_user_id = None
        player.is_verified = False
        player.verified_at = None
        await asyncio.to_thread(session.commit)
        session.close()
        await ctx.followup.send(f"✅ Unlinked **{player_name}** ({tag}) from {user.mention}.", ephemeral=True)

    @discord.slash_command(
        name="force_unlink",
        description="Unlink a CoC account from Discord without token verification (admin only)",
        default_member_permissions=discord.Permissions(administrator=True),
    )
    async def force_unlink(
        self,
        ctx: discord.ApplicationContext,
        tag: discord.Option(str, "Player tag, e.g. #ABC123", autocomplete=linked_tag_autocomplete),
    ):
        await ctx.defer(ephemeral=True)

        tag = tag.upper().replace("O", "0")
        if not tag.startswith("#"):
            tag = "#" + tag

        session = Session()
        player = await asyncio.to_thread(session.query(Player).filter_by(tag=tag).first)

        if not player or player.discord_user_id is None:
            session.close()
            await ctx.followup.send("❌ This account is not linked to any user.", ephemeral=True)
            return

        player_name = player.name
        player.discord_user_id = None
        player.is_verified = False
        player.verified_at = None
        await asyncio.to_thread(session.commit)
        session.close()
        await ctx.followup.send(f"✅ Unlinked **{player_name}** ({tag}).", ephemeral=True)

    @discord.slash_command(name="verify", description="Link and verify your Clash of Clans account")
    async def verify(
        self,
        ctx: discord.ApplicationContext,
        tag: discord.Option(str, "Your player tag, e.g. #ABC123"),
        api_token: discord.Option(str, "API token from in-game: Settings → More Settings → API Token"),
    ):
        await ctx.defer(ephemeral=True)

        tag = tag.upper().replace("O", "0")
        if not tag.startswith("#"):
            tag = "#" + tag

        valid = await verify_player_token(tag, api_token)
        if not valid:
            await ctx.followup.send(
                "❌ Invalid API token. Find it in-game: **Settings → More Settings → API Token**.",
                ephemeral=True,
            )
            return

        session = Session()
        try:
            player = await asyncio.to_thread(session.query(Player).filter_by(tag=tag).first)
            if not player:
                result = await add_player_to_db(tag, session)
                if not result["success"]:
                    await ctx.followup.send("❌ " + result["error"], ephemeral=True)
                    return
                player = await asyncio.to_thread(session.query(Player).filter_by(tag=result["tag"]).first)

            discord_user = await asyncio.to_thread(
                session.query(DiscordUser).filter_by(discord_id=str(ctx.author.id)).first
            )
            if not discord_user:
                discord_user = DiscordUser(discord_id=str(ctx.author.id))
                session.add(discord_user)
                await asyncio.to_thread(session.flush)

            if player.discord_user_id and player.discord_user_id != discord_user.id:
                await ctx.followup.send(
                    "❌ This account is already linked to a different Discord user.",
                    ephemeral=True,
                )
                return

            player.discord_user_id = discord_user.id
            player.is_verified = True
            player.verified_at = datetime.now(UTC)
            player_name = player.name
            await asyncio.to_thread(session.commit)
        finally:
            session.close()

        await ctx.followup.send(
            f"✅ **{player_name}** ({tag}) verified and linked to your account.",
            ephemeral=True,
        )

    @discord.slash_command(name="profile", description="Show CoC accounts linked to a Discord user")
    async def profile(
        self,
        ctx: discord.ApplicationContext,
        user: discord.Option(discord.Member, "Discord user (defaults to you)", required=False, default=None),
    ):
        await ctx.defer()
        target = user or ctx.author
        session = Session()

        discord_user = await asyncio.to_thread(session.query(DiscordUser).filter_by(discord_id=str(target.id)).first)
        if not discord_user or not discord_user.players:
            await ctx.followup.send(f"❌ {target.mention} has no linked CoC accounts.")
            session.close()
            return

        lines = [
            f"• [{p.name} ({p.tag})](https://link.clashofclans.com/en?action=OpenPlayerProfile&tag={p.tag.replace('#', '%23')})"
            for p in discord_user.players
        ]

        embed = discord.Embed(
            title=f"CoC Accounts — {target.display_name}",
            description="\n".join(lines),
            color=0x8B4513,
        )
        embed.set_thumbnail(url=target.display_avatar.url)
        session.close()
        await ctx.followup.send(embed=embed)


    @discord.slash_command(name="member_role_export", description="Export linked CoC accounts for members with a role to Excel")
    async def member_role_export(
        self,
        ctx: discord.ApplicationContext,
        role: discord.Option(discord.Role, "Discord role"),
    ):
        await ctx.defer()
        session = Session()

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = role.name[:31]

        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(fill_type="solid", fgColor="8B4513")
        headers = ["Discord ID", "Discord Name", "CoC Name", "CoC Tag", "TH Level", "Verified"]
        col_widths = [20, 25, 25, 14, 10, 10]

        for col, (h, w) in enumerate(zip(headers, col_widths), 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
            ws.column_dimensions[cell.column_letter].width = w

        discord_ids = [str(m.id) for m in role.members]
        players_by_discord_id = await asyncio.to_thread(_players_by_discord_id, session, discord_ids)
        session.close()

        row = 2
        unlinked = []
        for member in role.members:
            players = players_by_discord_id.get(str(member.id))
            if not players:
                unlinked.append(member.display_name)
                continue
            for player in players:
                ws.append([
                    str(member.id),
                    member.display_name,
                    player.name,
                    player.tag,
                    player.th_level,
                    "Yes" if player.is_verified else "No",
                ])
                row += 1

        if unlinked:
            ws_unlinked = wb.create_sheet("Unlinked")
            ws_unlinked.append(["Discord Name"])
            ws_unlinked.column_dimensions["A"].width = 25
            for name in unlinked:
                ws_unlinked.append([name])

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        now = datetime.now().strftime("%Y%m%d_%H%M")
        filename = f"members_{role.name}_{now}.xlsx".replace(" ", "_")
        await ctx.followup.send(
            f"📊 **{role.name}** — {row - 1} linked accounts, {len(unlinked)} unlinked",
            file=discord.File(buf, filename=filename),
        )

    @discord.slash_command(name="member_role", description="Show all linked CoC accounts for members with a role")
    async def member_role(
        self,
        ctx: discord.ApplicationContext,
        role: discord.Option(discord.Role, "Discord role"),
    ):
        await ctx.defer()
        session = Session()

        header = f"‎`{'DISCORD':<22} {'COC NAME':<18} {'TAG':<12} {'TH':>2}`"
        lines = [header]
        unlinked = []

        discord_ids = [str(m.id) for m in role.members]
        players_by_discord_id = await asyncio.to_thread(_players_by_discord_id, session, discord_ids)
        session.close()

        for member in role.members:
            players = players_by_discord_id.get(str(member.id))
            if not players:
                unlinked.append(member.display_name)
                continue
            discord_name = member.display_name[:22]
            for i, player in enumerate(players):
                name_col = discord_name if i == 0 else ""
                safe_player_name = player.name or "?"
                coc_name = "".join(c for c in safe_player_name if c.isascii() or "Ā" <= c <= "ɏ").strip()[:18] or safe_player_name[:18]
                th = str(player.th_level) if player.th_level else "—"
                lines.append(f"‎`{name_col:<22} {coc_name:<18} {player.tag:<12} {th:>2}`")

        embed = discord.Embed(title=f"👥 Members — {role.name}", color=0x8B4513)
        desc = "\n".join(lines)
        if len(desc) <= 4000:
            embed.description = desc
        else:
            block = header
            for line in lines[1:]:
                if len(block) + len(line) + 1 > 1024:
                    embed.add_field(name="", value=block, inline=False)
                    block = line
                else:
                    block += "\n" + line
            if block:
                embed.add_field(name="", value=block, inline=False)

        embed.set_footer(text=f"{len(role.members) - len(unlinked)} linked · {len(unlinked)} unlinked")
        await ctx.followup.send(embed=embed)

        if unlinked:
            await ctx.followup.send(
                "⚠️ No linked accounts: " + ", ".join(unlinked),
                ephemeral=True,
            )


    @discord.slash_command(name="verified_role", description="Show accounts and verification status for members with a role")
    async def verified_role(
        self,
        ctx: discord.ApplicationContext,
        role: discord.Option(discord.Role, "Discord role"),
    ):
        await ctx.defer()
        session = Session()

        header = f"‎`{'DISCORD':<22} {'COC NAME':<18} {'TAG':<12} {'VER'}`"
        lines = [header]
        unlinked = []

        discord_ids = [str(m.id) for m in role.members]
        players_by_discord_id = await asyncio.to_thread(_players_by_discord_id, session, discord_ids)
        session.close()

        for member in role.members:
            players = players_by_discord_id.get(str(member.id))
            if not players:
                unlinked.append(member)
                continue
            discord_name = member.display_name[:22]
            for i, player in enumerate(players):
                name_col = discord_name if i == 0 else ""
                safe_player_name = player.name or "?"
                coc_name = "".join(c for c in safe_player_name if c.isascii() or "Ā" <= c <= "ɏ").strip()[:18] or safe_player_name[:18]
                ver = "✅" if player.is_verified else "❌"
                lines.append(f"‎`{name_col:<22} {coc_name:<18} {player.tag:<12} {ver}`")

        linked_count = len(role.members) - len(unlinked)
        footer_text = f"{linked_count} with accounts · {len(unlinked)} unlinked"

        block = header
        chunks = []
        for line in lines[1:]:
            if len(block) + len(line) + 1 > 3800:
                chunks.append(block)
                block = header + "\n" + line
            else:
                block += "\n" + line
        chunks.append(block)

        for i, chunk in enumerate(chunks):
            embed = discord.Embed(color=0x8B4513)
            if i == 0:
                embed.title = f"🔐 Verification — {role.name}"
            if i == len(chunks) - 1:
                embed.set_footer(text=footer_text)
            embed.description = chunk
            await ctx.followup.send(embed=embed)

        if unlinked:
            unlinked_lines = [f"<@{m.id}> (`{m.id}`)" for m in unlinked]
            for chunk_start in range(0, len(unlinked_lines), 30):
                chunk = "\n".join(unlinked_lines[chunk_start:chunk_start + 30])
                prefix = "⚠️ **No linked accounts:**\n" if chunk_start == 0 else ""
                await ctx.followup.send(f"{prefix}{chunk}")


def setup(bot: discord.Bot):
    bot.add_cog(LinkCog(bot))
